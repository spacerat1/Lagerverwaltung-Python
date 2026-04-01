import os
import ctypes
import sqlite3
import pandas as pd
import re
import datetime
import win32com.client
import openpyxl as xl
from collections import defaultdict

from PyQt6.QtWidgets import (
    QFileDialog, QMenu, QApplication, QTreeWidgetItem, QLineEdit
)
from PyQt6.QtGui import QCursor
from PyQt6.QtCore import Qt

import application


# ── Hilfsfunktion: Widget aktivieren / deaktivieren ─────────────────
def set_widget_status(enabled: list, disabled: list, show:list = None, hide:list = None) -> None:
    
    for widget in disabled:
        if widget is not None:
            if isinstance(widget, QLineEdit):
                widget.clear()
            widget.setEnabled(False)

    for widget in enabled:
        if widget is not None:
            widget.setEnabled(True)    
    
    if show:
        for widget in show:
            widget.show()
    if hide:
        for widget in hide:
            widget.hide()


# ── Hilfsfunktion: alle Signal-Verbindungen der Filter-Widgets lösen ─
def unbind_all_widgets(app: application.App) -> None:
    """Trennt alle zuvor verbundenen Signale der Filter-Widgets."""
    for widget in (app.matnr_entry, app.sm_entry, app.posnr_entry):
        try:
            widget.returnPressed.disconnect()
        except TypeError:
            pass
        try:
            widget.textChanged.disconnect()
        except TypeError:
            pass
    try:
        app.combobox_loeschen.currentIndexChanged.disconnect()
    except TypeError:
        pass


# ── Datenbank öffnen / Pfad ändern ──────────────────────────────────
def open_db() -> str:
    settings = f"{os.environ['USERPROFILE']}\\Documents\\Lagerverwaltung_settings.txt"
    if not os.path.exists(settings):
        with open(settings, 'w'):
            pass
    with open(settings, 'r') as file:
        path_to_db = file.read()
    if not os.path.exists(path_to_db):
        ctypes.windll.user32.MessageBoxW(
            0,
            "Die Datenbank 'Lagerverwaltung_Datanbank.db' konnte nicht gefunden werden.\n"
            "Bitte im nächsten Fenster die Datenbank auswählen",
            "Pfad zur Datenbank suchen...",
            64
        )
        path_to_db = change_db_path()
    return path_to_db


def change_db_path(app: application.App = None) -> str:
    settings = f"{os.environ['USERPROFILE']}\\Documents\\Lagerverwaltung_settings.txt"
    path_to_db, _ = QFileDialog.getOpenFileName(
        None,
        'Datenbank auswählen',
        'Lagerverwaltung_Datenbank.db',
        'Datenbank (*.db);;Alle Dateien (*.*)'
    )
    if not path_to_db:
        return ''
    with open(settings, 'w') as file:
        file.write(path_to_db)
    if app:
        app.path_to_db = path_to_db
        color = '#228B22' if 'Service-Center' in path_to_db else '#FF4500'
        app.label_top_path_to_db.setText(f'Pfad zur Datenbank: {path_to_db}')
        app.label_top_path_to_db.setStyleSheet(
            f"color:{color};background:#000000;font:10pt Verdana; border: 1px solid {color}")
        app.connection.close()
        app.connection = sqlite3.connect(path_to_db)
        app.connection.row_factory = sqlite3.Row
        app.cursor = app.connection.cursor()
        # Ruft den zuletzt aktiven Button-Befehl erneut auf
        app.disabled_button.click()
    return path_to_db


# ── Kontextmenü (Rechtsklick) ────────────────────────────────────────
def show_context_menu(pos, app: application.App) -> None:
    selections = app.output_listbox.selectedItems()
    if not selections:
        return
    # Spaltenheader-Texte ermitteln
    col_count = app.output_listbox.columnCount()
    columns = [app.output_listbox.headerItem().text(c) for c in range(col_count)]

    menu = QMenu(app)
    menu.setStyleSheet(
        "QMenu{background:#000000;color:#228B22;font:10pt Verdana;}"
        "QMenu::item:selected{background:#228B22;color:#000000;}"
    )
    if 'MatNr.' in columns:
        menu.addAction('Kopieren: Materialnummer', lambda: copy_matnr(app))
    if 'SM Nummer / ID' in columns or 'SM Nummer' in columns:
        menu.addAction('Kopieren: SM Nummer', lambda: copy_sm_nr(app))
    if 'Bezeichnung' in columns:
        menu.addAction('Kopieren: Bezeichnung', lambda: copy_name(app))
    menu.addAction('Kopieren: Ganze Zeile', lambda: copy_line(app))
    menu.exec(QCursor.pos())


# ── Kopier-Hilfsfunktionen ───────────────────────────────────────────
def _col_index(app: application.App, col_name: str) -> int:
    """Gibt den Spaltenindex anhand des Header-Textes zurück, oder -1."""
    for c in range(app.output_listbox.columnCount()):
        if app.output_listbox.headerItem().text(c) == col_name:
            return c
    return -1


def copy_matnr(app: application.App) -> None:
    col = _col_index(app, 'MatNr.')
    if col < 0:
        return
    values = [item.text(col) for item in app.output_listbox.selectedItems()
              if item.text(col)]
    QApplication.clipboard().setText('\n'.join(values))


def copy_sm_nr(app: application.App) -> None:
    col = _col_index(app, 'SM Nummer / ID')
    if col < 0:
        col = _col_index(app, 'SM Nummer')
    if col < 0:
        return
    values = [item.text(col) for item in app.output_listbox.selectedItems()
              if item.text(col)]
    QApplication.clipboard().setText('\n'.join(values))


def copy_name(app: application.App) -> None:
    col = _col_index(app, 'Bezeichnung')
    if col < 0:
        return
    values = [item.text(col) for item in app.output_listbox.selectedItems()
              if item.text(col)]
    QApplication.clipboard().setText('\n'.join(values))


def copy_line(app: application.App) -> None:
    col_count = app.output_listbox.columnCount()
    lines = []
    for item in app.output_listbox.selectedItems():
        row_values = [item.text(c) for c in range(col_count)]
        lines.append('\t'.join(row_values))
    QApplication.clipboard().setText('\n'.join(lines))


# ── QTreeWidget-Hilfsfunktionen ──────────────────────────────────────
def _clear_tree(app: application.App) -> None:
    app.output_listbox.clear()


def _setup_columns(app: application.App, columns: list[str]) -> None:
    """Setzt Spalten-Header und -Breiten im QTreeWidget."""
    visible = [c for c in columns if c != 'LAST_COLUMN']
    app.output_listbox.setColumnCount(len(visible) + 1)  # +1 für Baum-Spalte (index 0)
    # Spalte 0 = Baum/Parent-Beschriftung (kein Daten-Header)
    app.output_listbox.headerItem().setText(0, '')
    for i, col in enumerate(visible, start=1):
        app.output_listbox.headerItem().setText(i, col)
        width, _ = app.columns_dict.get(col, (100, Qt.AlignmentFlag.AlignLeft))
        app.output_listbox.setColumnWidth(i, width)
    app.output_listbox.setColumnWidth(0, 220)  # Baum-Spalte


def _insert_parent(app: application.App, text: str,
                   tag: str = 'green') -> QTreeWidgetItem:
    """Fügt eine Parent-Zeile (Gruppe) ins QTreeWidget ein."""
    parent = QTreeWidgetItem(app.output_listbox)
    parent.setText(0, text)
    parent.setExpanded(True)
    app.apply_tag(parent, tag)
    return parent


def _insert_child(parent: QTreeWidgetItem, values: tuple | list,
                  app: application.App, tag: str = '') -> QTreeWidgetItem:
    """Fügt eine Kind-Zeile unter einem Parent ein."""
    child = QTreeWidgetItem(parent)
    # Werte ab Spalte 1 eintragen (Spalte 0 = Baum-Text, bleibt leer)
    for i, val in enumerate(values, start=1):
        child.setText(i, str(val) if val is not None else '')
        child.setExpanded(True)
    if tag:
        app.apply_tag(child, tag)
    return child


# ── Kritisches Material anzeigen ─────────────────────────────────────
def show_critical_material(app: application.App) -> None:
    
    unbind_all_widgets(app)
    app.disabled_button.setEnabled(True)
    app.disabled_button = app.button_krit_mat
    app.button_krit_mat.setEnabled(False)

    enabled  = []
    disabled = [app.sm_entry, app.posnr_entry, app.matnr_entry]
    show     = [app.bestellt_button, app.bestellt_label]
    hide     = [app.delete_button, app.print_button, app.combobox_loeschen]
    
    set_widget_status(enabled, disabled, show, hide)
    

    cursor = app.cursor
    ingoing_dict           = defaultdict(int)
    outgoing_dict          = defaultdict(int)
    already_ordered_dict   = defaultdict(int)
    date_dict              = defaultdict(str)
    mengen_dict            = defaultdict(str)

    ingoing              = cursor.execute("SELECT * FROM Wareneingang").fetchall()
    outgoing             = cursor.execute("SELECT * FROM Warenausgang").fetchall()
    outgoing_small       = cursor.execute(
        "SELECT * FROM Warenausgang_Kleinstmaterial_ohne_SM_BEZUG").fetchall()
    all_materials        = cursor.execute(
        "SELECT * FROM Standardmaterial UNION SELECT * FROM Kleinstmaterial").fetchall()

    for row in ingoing:
        ingoing_dict[row['MatNr']] += row['Menge']
    for row in outgoing:
        menge = row['Warenausgangsmenge'] if row['Warenausgangsmenge'] != 0 else row['Bedarfsmenge']
        if row['PosTyp'] == 8:
            outgoing_dict[row['MatNr']] -= menge
        elif row['PosTyp'] == 9:
            outgoing_dict[row['MatNr']] += menge
    for row in outgoing_small:
        outgoing_dict[row['MatNr']] += row['Menge']
    for row in all_materials:
        already_ordered_dict[row['MatNr']] = 'nein' if row['bestellt'] == 0 else 'ja'
        date_dict[row['MatNr']]  = row['Datum'] if row['bestellt'] else ''
        menge = row['Menge'] if row['Menge'] else ''
        mengen_dict[row['MatNr']] = menge if row['bestellt'] else ''

    small_material_output    = []
    standard_material_output = []
    for matnr, booked in sorted(outgoing_dict.items()):
        ingoing_mat    = ingoing_dict.get(matnr, 0)
        correction_sum = app.correction_dict.get(matnr, 0)
        bestand        = ingoing_mat - booked + correction_sum
        if bestand <= app.threshhold_dict[matnr] and matnr not in app.deprecated_dict:
            output = [matnr, app.materialnames_dict[matnr], bestand,
                      app.units_dict[matnr], app.recommended_amount_dict[matnr],
                      already_ordered_dict[matnr], mengen_dict[matnr], date_dict[matnr]]
            if matnr in app.standard_materials:
                standard_material_output.append(output)
            elif matnr in app.small_materials:
                small_material_output.append(output)

    columns = ['MatNr.', 'Bezeichnung', 'Bestand', 'Einheit',
               'empfohlene Menge', 'bestellt', 'Menge ', 'Datum', 'LAST_COLUMN']
    _clear_tree(app)
    _setup_columns(app, columns)

    if standard_material_output:
        parent = _insert_parent(app, 'Standardmaterial')
        for entry in standard_material_output:
            tag = 'red_font' if entry[2] <= 0 else 'green_font'
            _insert_child(parent, entry, app, tag)
    if small_material_output:
        parent = _insert_parent(app, 'Kleinstmaterial')
        for entry in small_material_output:
            _insert_child(parent, entry, app)
    if not standard_material_output and not small_material_output:
        item = QTreeWidgetItem(app.output_listbox)
        item.setText(2, 'Sieht gut aus, wir haben alles. :)')


# ── Bestellstatus umschalten ─────────────────────────────────────────
def toggle_ordered_status(app: application.App) -> None:
    connection = app.connection
    cursor     = app.cursor

    selections = app.output_listbox.selectedItems()
    mat_numbers, mat_names = [], []
    for item in selections:
        values = [item.text(c) for c in range(app.output_listbox.columnCount())]
        if not any(values):
            continue
        mat_numbers.append(values[1])   # Spalte 1 = MatNr.
        mat_names.append(values[2])     # Spalte 2 = Bezeichnung

    for mat_number, mat_name in zip(mat_numbers, mat_names):
        cursor.execute('''SELECT bestellt FROM Standardmaterial WHERE MatNr = ?
                          UNION SELECT bestellt FROM Kleinstmaterial WHERE MatNr = ?''',
                       (mat_number, mat_number))
        status = cursor.fetchone()[0]
        status = not status
        if not status:
            cursor.execute('UPDATE Standardmaterial SET bestellt = ? WHERE MatNr = ?',
                           (status, mat_number))
            cursor.execute('UPDATE Kleinstmaterial SET bestellt = ? WHERE MatNr = ?',
                           (status, mat_number))
        else:
            datum = datetime.datetime.strftime(datetime.datetime.now(), r'%d.%m.%Y %H:%M')
            menge = app.InputBox(mat_number, mat_name)
            cursor.execute('''UPDATE Standardmaterial
                              SET bestellt = ?, Datum = ?, Menge = ?
                              WHERE MatNr = ?''', (status, datum, menge, mat_number))
            cursor.execute('''UPDATE Kleinstmaterial
                              SET bestellt = ?, Datum = ?, Menge = ?
                              WHERE MatNr = ?''', (status, datum, menge, mat_number))
    connection.commit()
    show_critical_material(app)


# ── Bundle-Hilfsfunktionen ───────────────────────────────────────────
#
# deprecated - will be removed in future release
#
# def get_bundles_amount(matnr: int, app) -> int:
#     cursor       = app.cursor
#     ingoing_dict  = defaultdict(int)
#     outgoing_dict = defaultdict(int)
#     ingoing = cursor.execute("SELECT * FROM Wareneingang WHERE MatNr = ?", (matnr,)).fetchall()
#     outgoing = cursor.execute("SELECT * FROM Warenausgang WHERE MatNr = ?", (matnr,)).fetchall()
#     small_material = cursor.execute("SELECT * FROM Warenausgang_Kleinstmaterial_ohne_SM_BEZUG").fetchall()
#     for row in ingoing:
#         ingoing_dict[row['MatNr']] += row['Menge']
#     for row in outgoing:
#         menge = row['Warenausgangsmenge'] if row['Warenausgangsmenge'] != 0 else row['Bedarfsmenge']
#         if row['PosTyp'] == 8:
#             outgoing_dict[row['MatNr']] -= menge
#         elif row['PosTyp'] == 9:
#             outgoing_dict[row['MatNr']] += menge
#     for row in small_material:
#         outgoing_dict[row['MatNr']] += row['Menge']

#     bestand = 0
#     for matnr, menge in ingoing_dict.items():
#         bestand = menge - outgoing_dict.get(matnr, 0) + app.correction_dict.get(matnr, 0)
#     return bestand



# ── Bestand anzeigen ─────────────────────────────────────────────────
def show_stock(app: application.App) -> None:
    unbind_all_widgets(app)
    app.disabled_button.setEnabled(True)
    app.disabled_button = app.button_bestand
    app.button_bestand.setEnabled(False)

    enabled  = [app.matnr_entry]
    disabled = [app.posnr_entry, app.sm_entry]
    hide     = [app.bestellt_button,
                app.bestellt_label, app.delete_button, app.print_button,
                app.combobox_loeschen]

    set_widget_status(enabled, disabled, hide = hide)
    

    app.matnr_entry.returnPressed.connect(lambda: show_stock(app))
    app.matnr_entry.textChanged.connect(lambda: show_stock(app))
    app.matnr_entry.setFocus()

    cursor = app.cursor
    matnr  = app.matnr_entry.text()
    
    ingoing  = cursor.execute(
        "SELECT * FROM Wareneingang WHERE MatNr LIKE ?", (f'%{matnr}%',)).fetchall()
    outgoing = cursor.execute(
        "SELECT * FROM Warenausgang WHERE MatNr LIKE ?", (f'%{matnr}%',)).fetchall()
    bundles = cursor.execute(
        "SELECT * FROM Bundles").fetchall()
    bundles_in = cursor.execute(
        "SELECT * FROM Wareneingang WHERE MatNr in (SELECT MatNr FROM Bundles)").fetchall()
    bundles_out = cursor.execute(
        "SELECT * FROM Warenausgang WHERE MatNr in (SELECT MatNr FROM Bundles)").fetchall()
    
    ingoing_dict  = defaultdict(int)
    outgoing_dict = defaultdict(int)
    bundles_in_dict = defaultdict(int)
    bundles_out_dict = defaultdict(int)
    bundles_total_dict = defaultdict(list)

    for row in ingoing:
        ingoing_dict[row['MatNr']] += row['Menge']
    for row in bundles_in:
        bundles_in_dict[row['MatNr']] += row['Menge']
    for row in outgoing:
        menge = row['Warenausgangsmenge'] if row['Warenausgangsmenge'] != 0 else row['Bedarfsmenge']
        if row['PosTyp'] == 8:
            outgoing_dict[row['MatNr']] -= menge
        elif row['PosTyp'] == 9:
            outgoing_dict[row['MatNr']] += menge
    for row in bundles_out:
        menge = row['Warenausgangsmenge'] if row['Warenausgangsmenge'] != 0 else row['Bedarfsmenge']
        if row['PosTyp'] == 8:
            bundles_out_dict[row['MatNr']] -= menge
        elif row['PosTyp'] == 9:
            bundles_out_dict[row['MatNr']] += menge        

    
    for row in bundles:
        menge = bundles_in_dict.get(row['MatNr'], 0) - bundles_out_dict.get(row['MatNr'], 0) + app.correction_dict.get(row['MatNr'], 0)
        bundles_total_dict[row['MatNr']] = [row['Bezeichnung'], eval(row['Packungsinhalt']), menge]
    
    for row in cursor.execute(
            "SELECT * FROM Warenausgang_Kleinstmaterial_ohne_SM_BEZUG").fetchall():
        outgoing_dict[row['MatNr']] += row['Menge']


    standardmaterial_list = []
    small_material_list   = []
    for matnr_key, menge in ingoing_dict.items():
        bezeichnung       = app.materialnames_dict[matnr_key]
        einheit           = app.units_dict[matnr_key]
        bemerkung         = ''
        bundle_correction = 0 # Das Material aus Bundles wird den Einzelpositionen hinzugerechnet
        for bundle, data in bundles_total_dict.items():
            _, content, bundle_menge = data
            if matnr_key in content:
                bundle_correction = bundle_menge
                if bundle_correction:
                    if not bemerkung:
                        bemerkung += f'davon {bundle_menge} {einheit} aus Bundle {bundle}'
                    else:
                        bemerkung += f", {bundle_menge} {einheit} aus {bundle}"
        bestand = menge - outgoing_dict.get(matnr_key, 0) + app.correction_dict.get(matnr_key, 0) + bundle_correction
        
        if matnr_key in app.standard_materials:
            standardmaterial_list.append((matnr_key, bezeichnung, bestand, einheit, bemerkung))
        elif matnr_key in app.small_materials:
            small_material_list.append((matnr_key, bezeichnung, bestand, einheit, bemerkung))

    columns = ['MatNr.', 'Bezeichnung', 'Bestand', 'Einheit', 'Bemerkungen', 'LAST_COLUMN']
    _clear_tree(app)
    _setup_columns(app, columns)

    if standardmaterial_list:
        parent = _insert_parent(app, 'Standardmaterial')
        for entry in sorted(standardmaterial_list, key=lambda x: int(x[0])):
            _insert_child(parent, entry, app)
    if small_material_list:
        parent = _insert_parent(app, 'Kleinstmaterial')
        for entry in sorted(small_material_list, key=lambda x: int(x[0])):
            _insert_child(parent, entry, app)
    if not standardmaterial_list and not small_material_list:
        item = QTreeWidgetItem(app.output_listbox)
        item.setText(2, 'Keine Daten gefunden. Bitte Filter prüfen')


# ── Wareneingang anzeigen ────────────────────────────────────────────
def show_ingoing_material(app: application.App) -> None:
    
    unbind_all_widgets(app)
    app.disabled_button.setEnabled(True)
    app.disabled_button = app.button_wareneingang
    app.button_wareneingang.setEnabled(False)

    enabled  = [app.matnr_entry]
    disabled = [app.sm_entry, app.posnr_entry, ]
    hide     = [app.bestellt_label,
                app.bestellt_button, app.delete_button, app.print_button,
                app.combobox_loeschen]

    set_widget_status(enabled, disabled, hide = hide)
    
    app.matnr_entry.setFocus()
    app.matnr_entry.returnPressed.connect(lambda: show_ingoing_material(app))
    app.matnr_entry.textChanged.connect(lambda: show_ingoing_material(app))

    cursor = app.cursor
    matnr  = app.matnr_entry.text()

    selection  = cursor.execute(
        "SELECT * FROM Wareneingang WHERE MatNr LIKE ?", (f'%{matnr}%',)).fetchall()
    correction = cursor.execute(
        "SELECT * FROM Jahresinventur_Korrekturdaten WHERE MatNr LIKE ? AND Menge > 0",
        (f'%{matnr}%',)).fetchall()

    columns = ['ID', 'MatNr.', 'Bezeichnung', 'Menge', 'Einheit', 'Datum', 'LAST_COLUMN']
    _clear_tree(app)
    _setup_columns(app, columns)

    if selection:
        parent = _insert_parent(app, 'Standardlieferung')
        for entry in reversed(selection):
            date = datetime.datetime.strftime(
                datetime.datetime.strptime(entry['Datum'], r"%Y-%m-%d %H:%M:%S"),
                r"%d.%m.%Y %H:%M:%S")
            _insert_child(parent, (entry['ID'], entry['MatNr'], entry['Bezeichnung'],
                                   entry['Menge'], app.units_dict[entry['MatNr']], date), app)
    if correction:
        parent = _insert_parent(app, 'Inventurkorrektur')
        for entry in correction:
            date = datetime.datetime.strftime(
                datetime.datetime.strptime(entry['Datum'], r"%Y-%m-%d %H:%M:%S"),
                r"%d.%m.%Y %H:%M:%S")
            _insert_child(parent, (entry['ID'], entry['MatNr'], entry['Bezeichnung'],
                                   entry['Menge'], app.units_dict[entry['MatNr']], date), app)
    if not selection:
        item = QTreeWidgetItem(app.output_listbox)
        item.setText(3, 'Keine Daten gefunden. Bitte Filter prüfen.')


# ── Warenausgang anzeigen ────────────────────────────────────────────
def show_outgoing_material(app: application.App) -> None:
    unbind_all_widgets(app)
    app.disabled_button.setEnabled(True)
    app.disabled_button = app.button_warenausgang
    app.button_warenausgang.setEnabled(False)

    enabled  = [app.sm_entry, app.matnr_entry]
    disabled = [app.posnr_entry]
    hide     = [app.bestellt_label, app.bestellt_button,
                app.delete_button, app.print_button, app.combobox_loeschen]
    
    set_widget_status(enabled, disabled, hide = hide)
    

    app.sm_entry.returnPressed.connect(lambda: show_outgoing_material(app))
    app.sm_entry.textChanged.connect(lambda: show_outgoing_material(app))
    app.matnr_entry.returnPressed.connect(lambda: show_outgoing_material(app))
    app.matnr_entry.textChanged.connect(lambda: show_outgoing_material(app))
    focused = QApplication.focusWidget()
    if focused not in (app.sm_entry, app.matnr_entry):
        app.matnr_entry.setFocus()

    cursor = app.cursor
    matnr  = app.matnr_entry.text()
    smnr   = app.sm_entry.text()

    selection_with_sm = cursor.execute(
        "SELECT * FROM Warenausgang WHERE MatNr LIKE ? AND SM_Nummer LIKE ? AND PosTyp = 9",
        (f'%{matnr}%', f'%{smnr}%')).fetchall()
    selection_without_sm = cursor.execute(
        "SELECT * FROM Warenausgang_Kleinstmaterial_ohne_SM_Bezug WHERE MatNr LIKE ?",
        (f'%{matnr}%',)).fetchall()
    correction = cursor.execute(
        "SELECT * FROM Jahresinventur_Korrekturdaten WHERE MatNr LIKE ? AND Menge < 0",
        (f'%{matnr}%',)).fetchall()

    columns = ['Nr.', 'SM Nummer / ID', 'Position', 'Pos.Typ', 'MatNr.',
               'Bezeichnung', 'Menge', 'Einheit', 'Lieferschein',
               'SD Beleg', 'Materialbeleg', 'LAST_COLUMN']
    _clear_tree(app)
    _setup_columns(app, columns)
    app.output_listbox.setColumnWidth(0, 280)

    if selection_with_sm:
        selection_with_sm = sorted(selection_with_sm, key=lambda x: x['SM_Nummer'])
        parent = _insert_parent(app, 'Warenausgang mit SM Bezug')
        for number, entry in enumerate(selection_with_sm, start=1):
            _insert_child(parent, (number, entry['SM_Nummer'], entry['Position'],
                                   entry['PosTyp'], entry['MatNr'], entry['Bezeichnung'],
                                   entry['Bedarfsmenge'], app.units_dict[entry['MatNr']],
                                   entry['Lieferschein'][:-2], entry['SD_Beleg'][:-2],
                                   entry['Materialbeleg'][:-2]), app)
    if selection_without_sm and not smnr:
        parent = _insert_parent(app, 'Warenausgang ohne SM Bezug')
        for number, entry in enumerate(selection_without_sm,
                                       start=len(selection_with_sm) + 1):
            _insert_child(parent, (number, entry['ID'], '', '', entry['MatNr'],
                                   entry['Bezeichnung'], entry['Menge'],
                                   app.units_dict[entry['MatNr']]), app)
    if correction:
        parent = _insert_parent(app, 'Inventurkorrektur')
        start = len(selection_with_sm) + (len(selection_without_sm) if not smnr else 0)
        for number, entry in enumerate(correction, start=start + 1):
            date = datetime.datetime.strftime(
                datetime.datetime.strptime(entry['Datum'], r"%Y-%m-%d %H:%M:%S"),
                r"%d.%m.%Y %H:%M:%S")
            _insert_child(parent, (number, entry['ID'], '', '', entry['MatNr'],
                                   entry['Bezeichnung'], abs(entry['Menge']),
                                   app.units_dict[entry['MatNr']], date), app)
    if not selection_with_sm and not selection_without_sm and not correction:
        item = QTreeWidgetItem(app.output_listbox)
        item.setText(6, 'Keine Daten gefunden. Bitte Filter prüfen')


# ── Material für SM-Auftrag anzeigen ─────────────────────────────────
def show_material_for_order(app: application.App) -> None:
    
    unbind_all_widgets(app)
    
    app.disabled_button.setEnabled(True)
    app.disabled_button = app.button_sm_auftrag
    app.button_sm_auftrag.setEnabled(False)

    enabled  = [app.sm_entry]
    disabled = [app.posnr_entry, 
                app.matnr_entry
                ]
    show     = [app.print_button]
    hide     = [app.bestellt_label,
                app.bestellt_button, 
                app.delete_button, 
                app.combobox_loeschen
                ]
    set_widget_status(enabled, disabled, show = show, hide = hide)
    
    

    app.sm_entry.returnPressed.connect(lambda: show_material_for_order(app))
    app.sm_entry.textChanged.connect(lambda: show_material_for_order(app))
    app.sm_entry.setFocus()

    cursor = app.cursor
    smnr   = app.sm_entry.text()
    selection          = cursor.execute(
        "SELECT * FROM Warenausgabe_Comline WHERE SM_Nummer LIKE ?",
        (f'%{smnr}%',)).fetchall()
    selection_addresses= cursor.execute(
        "SELECT * FROM Adresszuordnung WHERE SM_Nummer LIKE ?",
        (f'%{smnr}%',)).fetchall()
    bundles      = cursor.execute("SELECT * FROM Bundles").fetchall()
    bundles_dict = {}
 # ---------------------- Hier bundles dict erzeugen, nicht jede Position einzeln ermitteln  
    for row in bundles:
        bundles_dict[row['MatNr']] = [row['Bezeichnung'], eval(row['Packungsinhalt'])]
 # --------------------------------------------------------------------------------------------   
    standard_material = []
    small_material    = []
    telekom_material  = []
    for row in selection:
        values = (row['SM_Nummer'], row['MatNr'], row['Bezeichnung'],
                  row['Bedarfsmenge'], app.units_dict[row['MatNr']] or 'n/a')
        if row['MatNr'] in app.standard_materials:
            standard_material.append(values)
        elif row['MatNr'] in app.small_materials:
            small_material.append(values)
        else:
            telekom_material.append(values)
    address_values = [(r['SM_Nummer'], r['VPSZ'], r['Adresse']) for r in selection_addresses]

    columns = ['SM Nummer', 'MatNr.', 'Bezeichnung', 'Menge', 'Einheit', 'LAST_COLUMN']
    _clear_tree(app)
    _setup_columns(app, columns)
    app.output_listbox.setColumnWidth(0, 250)

    if not selection:
        item = QTreeWidgetItem(app.output_listbox)
        item.setText(3, 'Keine Daten gefunden. Bitte Filter prüfen.')
        return

    if address_values:
        parent = _insert_parent(app, 'Adressen')
        for entry in address_values:
            _insert_child(parent, entry, app)

    if standard_material:
        parent = _insert_parent(app, 'Standardmaterial')
        for entry in standard_material:
            bundle_nr = entry[1]
            amount    = entry[3]
            if bundle_nr in bundles_dict:
                bundle_item = _insert_child(parent, entry, app, 'bundle_head')
                bundle_name, contents = bundles_dict[bundle_nr]
                for matnr in contents:
                    name = app.materialnames_dict[matnr]
                    unit = app.units_dict[matnr]
                    _insert_child(bundle_item, ('', matnr, name, amount, unit), app, 'bundle')
            else:
                _insert_child(parent, entry, app)

    if small_material:
        parent = _insert_parent(app, 'Kleinstmaterial')
        for entry in small_material:
            _insert_child(parent, entry, app)

    if telekom_material:
        parent = _insert_parent(app, 'Telekommaterial')
        for entry in telekom_material:
            _insert_child(parent, entry, app)


# ── Drucken ──────────────────────────────────────────────────────────
def print_screen(app: application.App) -> None:
    file_path = f"{os.getcwd()}\\Ausgabe.xlsx"
    if len(app.sm_entry.text()) < 4:
        ctypes.windll.user32.MessageBoxW(
            0,
            "Um Fehlausdrucke zu vermeiden müssen im Filter SM-Nummer \n"
            "mindestens 4 Zeichen eingetragen sein.",
            "Zu wenig Zeichen bei SM Nummer", 64)
        return
    answer = ctypes.windll.user32.MessageBoxW(
        0, "Soll der angezeigte Inhalt gedruckt werden?", "Drucken...", 68)
    if answer != 6:
        return

    standard_material = []
    small_material    = []
    telekom_material  = []
    address           = []
    col_count         = app.output_listbox.columnCount()

    def _get_values(item: QTreeWidgetItem) -> list[str]:
        return [item.text(c) for c in range(col_count)]

    root = app.output_listbox.invisibleRootItem()
    for i in range(root.childCount()):
        parent     = root.child(i)
        group_text = parent.text(0)
        for j in range(parent.childCount()):
            child  = parent.child(j)
            values = _get_values(child)[1:]
            output = '\t'.join(values)
            if group_text == 'Standardmaterial':
                standard_material.append(output)
            elif group_text == 'Kleinstmaterial':
                small_material.append(output)
            elif group_text == 'Adressen':
                address.append(output)
            else:
                telekom_material.append(output)
            # Sub-Children (Bundles)
            for k in range(child.childCount()):
                sub     = child.child(k)
                subvals = _get_values(sub)[1:]
                subout  = '\t'.join(subvals)
                if group_text == 'Standardmaterial':
                    standard_material.append(subout)
                elif group_text == 'Kleinstmaterial':
                    small_material.append(subout)
                elif group_text == 'Adressen':
                    address.append(subout)
                else:
                    telekom_material.append(subout)

    wb    = xl.Workbook()
    sheet = wb.active
    sheet.page_setup.orientation = 'landscape'
    sheet.page_setup.fitToPage   = True
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 16
    sheet.column_dimensions['C'].width = 44
    sheet.column_dimensions['D'].width = 7
    sheet.column_dimensions['E'].width = 7
    sheet.column_dimensions['F'].width = 3
    for x in range(7, 37):
        letter = xl.utils.cell.get_column_letter(x)
        sheet.column_dimensions[letter].width = 1 if x % 2 == 1 else 3

    start = 1
    if address:
        sheet.cell(column=1, row=start, value='Adresse')
        sheet.cell(column=1, row=start).font = xl.styles.Font(size=12, bold=True)
        start += 1
        for row, line in enumerate(address, start=start):
            for col, word in enumerate(line.split('\t'), start=1):
                sheet.cell(column=col, row=row, value=word)
        start += len(address) + 1
    if standard_material:
        sheet.cell(column=1, row=start, value='Standardmaterial')
        sheet.cell(column=1, row=start).font = xl.styles.Font(size=12, bold=True)
        start += 1
        for row, line in enumerate(standard_material, start=start):
            for col, word in enumerate(line.split('\t'), start=1):
                sheet.cell(column=col, row=row, value=word)
        start += len(standard_material) + 1
    if small_material:
        sheet.cell(column=1, row=start, value='Kleinstmaterial')
        sheet.cell(column=1, row=start).font = xl.styles.Font(size=12, bold=True)
        start += 1
        for row, line in enumerate(small_material, start=start):
            for col, word in enumerate(line.split('\t'), start=1):
                sheet.cell(column=col, row=row, value=word)
        start += len(small_material) + 1
    if telekom_material:
        sheet.cell(column=1, row=start, value='Telekommaterial')
        sheet.cell(column=1, row=start).font = xl.styles.Font(size=12, bold=True)
        start += 1
        for row, line in enumerate(telekom_material, start=start):
            for col, word in enumerate(line.split('\t'), start=1):
                sheet.cell(column=col, row=row, value=word)

    border = xl.styles.Side(border_style='thin', color='000000')
    bundle = False
    for idx, row in enumerate(sheet, start=1):
        if not row[1].value:
            continue
        if not sheet.cell(row=idx, column=5).value:
            for col in range(1, 4):
                sheet.cell(row=idx, column=col).font = xl.styles.Font(color="3282F6")
            continue
        if row[0].value:
            start_col = 1
            bundle    = False
        else:
            start_col = 2
            for col in range(1, 6):
                if not bundle:
                    sheet.cell(row=idx - 1, column=col).fill = xl.styles.PatternFill(
                        start_color="B6B6B6", fill_type="solid")
                sheet.cell(row=idx, column=col).fill = xl.styles.PatternFill(
                    start_color="E6E6E6", fill_type="solid")
            bundle = True
        for col in range(start_col, 6):
            sheet.cell(row=idx, column=col).border = xl.styles.Border(bottom=border)
        try:
            if (int(sheet.cell(row=idx, column=4).value) > 1 and
                    sheet.cell(row=idx, column=5).value.strip() in ('M', 'n/a')):
                for col in range(1, 6):
                    sheet.cell(row=idx, column=col).font = xl.styles.Font(
                        color="FF0000", bold=True)
        except TypeError, ValueError:
            pass
        if sheet.cell(row=idx, column=5).value.strip() in ('ST', 'SA', 'PAK'):
            amount = int(sheet.cell(row=idx, column=4).value) * 2
            if amount > 30:
                amount = 2
        else:
            amount = 2
        for col in range(amount):
            if col % 2 == 0:
                continue
            sheet.cell(row=idx, column=col + 7).border = xl.styles.Border(
                bottom=border, top=border, left=border, right=border)

    wb.save(file_path)
    try:
        excel_app = win32com.client.Dispatch('Excel.Application')
        excel_app.Visible = False
        wb_com = excel_app.Workbooks.Open(file_path)
        wb_com.PrintOut()
    finally:
        wb_com.Close(False)
        excel_app.Quit()
        os.remove(file_path)


# ── Warenausgang aus Excel buchen ────────────────────────────────────
def book_outgoing_from_excel_file(app: application.App) -> None:
    app.disabled_button.setEnabled(True)
    app.disabled_button = app.button_warenausgang_buchen

    enabled  = []
    disabled = [app.sm_entry, 
                app.posnr_entry, 
                app.matnr_entry,
                ]
    
    hide     = [app.delete_button, 
                app.print_button, 
                app.combobox_loeschen,
                app.bestellt_button, 
                app.bestellt_label]
    
    set_widget_status(enabled, disabled, hide = hide)

    cursor     = app.cursor
    connection = app.connection

    paths, _ = QFileDialog.getOpenFileNames(
        app, 'Bitte die EXPORT Datei aus PSL wählen – Mehrfachauswahl möglich...',
        '', 'Excel (*.xlsx);;Alle Dateien (*.*)')
    if not paths:
        show_critical_material(app)
        return

    all_rows = new_entries = no_booking = already_exists = 0

    for file in paths:
        excel_file             = pd.read_excel(file)
        valid_entries          = []
        absolute_valid_entries = []
        ausgabe_entries        = []
        valid_ausgabe_entries  = []

        kleinstmaterial  = [r[0] for r in cursor.execute(
            'SELECT MatNr FROM Kleinstmaterial').fetchall()]
        standardmaterial = [r[0] for r in cursor.execute(
            'SELECT MatNr FROM Standardmaterial').fetchall()]

        for _, line in excel_file.iterrows():
            if line['Positionstyp'] in (9, '9', 'N'):
                ausgabe_entries.append(line)
            if line['Material'] in kleinstmaterial:
                valid_entries.append(line)
                continue
            if (line['Material'] in standardmaterial and
                    (not pd.isnull(line['Lieferungen(ab/bis)']) or
                     line['Positionstyp'] in (8, '8'))):
                valid_entries.append(line)

        for entry in valid_entries:
            cursor.execute(
                f"SELECT * FROM Warenausgang WHERE SM_Nummer = {entry['Auftrag']}"
                f" AND Position = {entry['Stl.Position']}")
            if not cursor.fetchall():
                absolute_valid_entries.append(entry)

        for entry in ausgabe_entries:
            cursor.execute(
                f"SELECT * FROM Warenausgabe_Comline WHERE SM_Nummer = {entry['Auftrag']}"
                f" AND Position = {entry['Stl.Position']}")
            if not cursor.fetchall():
                valid_ausgabe_entries.append(entry)

        for line in absolute_valid_entries:
            entry = line.fillna(" ")
            cursor.execute('''
                INSERT INTO Warenausgang
                    (SM_Nummer, Position, PosTyp, MatNr, Bezeichnung, SD_Beleg,
                     Bedarfsmenge, Warenausgangsmenge, Umbuchungsmenge,
                     Lieferschein, Materialbeleg)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
                (entry['Auftrag'], entry['Stl.Position'], entry['Positionstyp'],
                 entry['Material'], entry['Materialkurztext'], entry['SD Beleg'],
                 entry['Bedarfsmenge'], entry['Warenausgangsmenge'],
                 entry['Umbuchungsmenge'], entry['Lieferschein'], entry['Materialbeleg']))

        for line in valid_ausgabe_entries:
            entry = line.fillna(" ")
            cursor.execute('''
                INSERT INTO Warenausgabe_Comline
                    (SM_Nummer, Position, PosTyp, MatNr, Bezeichnung, SD_Beleg,
                     Bedarfsmenge, Warenausgangsmenge, Umbuchungsmenge,
                     Lieferschein, Materialbeleg)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
                (entry['Auftrag'], entry['Stl.Position'], entry['Positionstyp'],
                 entry['Material'], entry['Materialkurztext'], entry['SD Beleg'],
                 entry['Bedarfsmenge'], entry['Warenausgangsmenge'],
                 entry['Umbuchungsmenge'], entry['Lieferschein'], entry['Materialbeleg']))

        connection.commit()
        rows_in_excel        = excel_file.shape[0]
        valid_count          = len(valid_entries)
        absolute_valid_count = len(absolute_valid_entries)
        all_rows             += rows_in_excel
        new_entries          += absolute_valid_count
        no_booking           += rows_in_excel - valid_count
        already_exists       += valid_count - absolute_valid_count

    ctypes.windll.user32.MessageBoxW(
        0,
        f"{new_entries} von {all_rows} Datensätze in Datenbank aufgenommen.\n"
        f"{no_booking} Einträge enthielten keine Bestellung.\n"
        f"{already_exists} Einträge waren bereits in der Datenbank enthalten.",
        'Ergebnis', 64)
    show_critical_material(app)


# ── Wareneingang buchen ──────────────────────────────────────────────
def book_ingoing_position(app: application.App) -> None:
    connection = app.connection
    cursor     = app.cursor

    app.disabled_button.setEnabled(True)
    app.disabled_button = app.button_wareneingang_buchen
    app.button_wareneingang_buchen.setEnabled(False)

    enabled  = []
    disabled = [app.sm_entry, 
                app.posnr_entry, 
                app.matnr_entry,
                ]
    
    hide     = [app.delete_button, 
                app.print_button, 
                app.combobox_loeschen,
                app.bestellt_button, 
                app.bestellt_label]
    
    set_widget_status(enabled, disabled, hide = hide)

    title     = 'Wareneingang buchen (Anlieferung CTDI)'
    selection = 'SELECT * FROM Kleinstmaterial UNION SELECT * FROM Standardmaterial'
    app.open_booking_window(title, selection)

    if app.user_closed_window:
        show_ingoing_material(app)
        return

    try:
        matnr = re.findall(r'\d{8}', app.ingoing_mat_string_var.currentText())[0]
    except IndexError:
        show_ingoing_material(app)
        return

    try:
        menge = int(app.ingoing_menge_var.text())
    except ValueError:
        menge = 0
    if menge == 0:
        show_ingoing_material(app)
        return

    material_query = ("SELECT * FROM Standardmaterial WHERE MatNr = ? "
                      "UNION SELECT * FROM Kleinstmaterial WHERE MatNr = ?")
    cursor.execute(material_query, (matnr, matnr))
    material    = cursor.fetchall()[0]
    bezeichnung = material['Bezeichnung']
    einheit     = material['Einheit']

    answer = ctypes.windll.user32.MessageBoxW(
        0,
        f"Soll folgendes Material gebucht werden?\n\n"
        f"{matnr}    '{bezeichnung}'       {menge} {einheit}",
        "Warenausgang buchen...", 68)
    if answer != 6:
        show_ingoing_material(app)
        return

    cursor.execute(
        'INSERT INTO Wareneingang (MatNr, Bezeichnung, Menge) VALUES (?,?,?)',
        (matnr, bezeichnung, menge))
    cursor.execute(
        f'UPDATE Standardmaterial SET bestellt = 0 WHERE MatNr = {matnr}')
    cursor.execute(
        f'UPDATE Kleinstmaterial SET bestellt = 0 WHERE MatNr = {matnr}')
    connection.commit()
    app.button_wareneingang_buchen.setEnabled(True)
    show_ingoing_material(app)


# ── Kleinstmaterial ohne SM Bezug buchen ─────────────────────────────
def book_outgoing_kleinstmaterial(app: application.App) -> None:
    connection = app.connection
    cursor     = app.cursor

    app.disabled_button.setEnabled(True)
    app.disabled_button = app.button_warenausgang_buchen_Kleinstmaterial
    app.button_warenausgang_buchen_Kleinstmaterial.setEnabled(False)

    enabled  = []
    disabled = [app.sm_entry, 
                app.posnr_entry, 
                app.matnr_entry,
                ]
    
    hide     = [app.delete_button, 
                app.print_button, 
                app.combobox_loeschen,
                app.bestellt_button, 
                app.bestellt_label]
    
    set_widget_status(enabled, disabled, hide = hide)

    title     = 'Kleinstmaterial ohne SM Bezug herausgeben'
    selection = 'SELECT * FROM Kleinstmaterial'
    app.open_booking_window(title, selection)

    if app.user_closed_window:
        show_critical_material(app)
        return

    try:
        matnr = re.findall(r'\d{8}', app.ingoing_mat_string_var.currentText())[0]
    except IndexError:
        show_critical_material(app)
        return

    try:
        menge = int(app.ingoing_menge_var.text())
    except ValueError:
        menge = 0
    if menge == 0:
        show_critical_material(app)
        return

    cursor.execute("SELECT * FROM Kleinstmaterial WHERE MatNr = ?", (matnr,))
    material    = cursor.fetchall()[0]
    bezeichnung = material['Bezeichnung']
    einheit     = material['Einheit']

    answer = ctypes.windll.user32.MessageBoxW(
        0,
        f"Soll folgendes Material gebucht werden?\n\n"
        f"{matnr}    '{bezeichnung}'       {menge} {einheit}",
        "Warenausgang buchen...", 68)
    if answer != 6:
        show_critical_material(app)
        return

    cursor.execute(
        'INSERT INTO Warenausgang_Kleinstmaterial_ohne_SM_Bezug '
        '(MatNr, Bezeichnung, Menge) VALUES (?,?,?)',
        (matnr, bezeichnung, menge))
    connection.commit()
    app.button_warenausgang_buchen_Kleinstmaterial.setEnabled(True)
    show_critical_material(app)


# ── Einträge aus Datenbank löschen ───────────────────────────────────
def filter_entries_to_delete(app: application.App) -> None:
    cursor = app.cursor

    app.disabled_button.setEnabled(True)
    app.disabled_button = app.button_loeschen
    app.button_loeschen.setEnabled(False)

    if not app.combobox_loeschen.currentText():
        app.combobox_loeschen.setCurrentIndex(0)

    enabled  = [app.sm_entry, 
                app.posnr_entry, 
                app.matnr_entry
                ]
    disabled = []
    enabled.extend(app.filter_dict[app.combobox_loeschen.currentText()][0])
    disabled.extend(app.filter_dict[app.combobox_loeschen.currentText()][1])
    show = [app.delete_button, 
            app.combobox_loeschen
            ]
    hide = [app.bestellt_label, 
            app.bestellt_button, 
            app.print_button
            ]
    
    set_widget_status(enabled, disabled, show = show, hide = hide)

    app.sm_entry.textChanged.connect(lambda: filter_entries_to_delete(app))
    app.posnr_entry.textChanged.connect(lambda: filter_entries_to_delete(app))
    app.matnr_entry.textChanged.connect(lambda: filter_entries_to_delete(app))
    app.combobox_loeschen.currentIndexChanged.connect(
        lambda: filter_entries_to_delete(app))

    sm     = app.sm_entry.text()
    posnr  = app.posnr_entry.text()
    matnr  = app.matnr_entry.text()
    table  = app.combobox_loeschen.currentText()

    execution_string = app.execution_dict[table][0]
    execute_values   = app.execution_dict[table][1]
    execute_values   = execute_values.replace('matnr', matnr).replace('sm', sm).replace('posnr', posnr)
    execution_tuple  = tuple([e for e in execute_values.split(',') if e])
    if not execution_tuple:
        execution_tuple = ('',)

    columns_info = cursor.execute(f"PRAGMA table_info ({table})").fetchall()
    columns      = [info[1] for info in columns_info]
    columns.append('LAST_COLUMN')

    _clear_tree(app)
    _setup_columns(app, columns)

    selection = cursor.execute(execution_string, execution_tuple).fetchall()
    for row in selection:
        values = list(row)
        item   = QTreeWidgetItem(app.output_listbox)
        for i, val in enumerate(values, start=1):
            item.setText(i, str(val) if val is not None else '')


def delete_selected_entries(app: application.App) -> None:
    selections = app.output_listbox.selectedItems()
    if not selections:
        ctypes.windll.user32.MessageBoxW(
            0,
            "Es sind keine Einträge markiert.\nNur markierte Einträge werden gelöscht.",
            "Markier was..", 64)
        return
    answer = ctypes.windll.user32.MessageBoxW(
        0,
        "Achtung!!\nDie markierten Einträge werden unwiderruflich gelöscht.\n"
        "Bist du sicher, dass du die Einträge löschen willst?",
        "Einträge löschen", 68)
    if answer != 6:
        return

    col_count = app.output_listbox.columnCount()
    headers   = [app.output_listbox.headerItem().text(c) for c in range(col_count)]

    def _get_col(item: QTreeWidgetItem, name: str) -> str:
        try:
            idx = headers.index(name)
            return item.text(idx)
        except ValueError:
            return ''

    for item in selections:
        matnr = _get_col(item, 'MatNr')
        sm    = _get_col(item, 'SM_Nummer')
        posnr = _get_col(item, 'ID') or _get_col(item, 'Position')

        execution_string = app.deletion_dict[app.combobox_loeschen.currentText()][0]
        execute_values   = app.deletion_dict[app.combobox_loeschen.currentText()][1]
        execute_values   = execute_values.replace('matnr', matnr).replace('sm', sm).replace('posnr', posnr)
        execution_tuple  = tuple([e for e in execute_values.split(',') if e])
        if not execution_tuple:
            execution_tuple = ('',)
        app.connection.execute(execution_string, execution_tuple)

    app.connection.commit()
    filter_entries_to_delete(app)


# ── Buchungsfenster bestätigen ───────────────────────────────────────
def confirm_user_input(app: application.App) -> None:
    app.user_closed_window = False
    app.ingoing_window.accept()   # QDialog schließen (statt quit/destroy)


# ── Datenbankpflege (unveränderte Logik, nur filedialog ersetzt) ──────
def add_kleinstmaterial(connection: sqlite3.Connection,
                        cursor: sqlite3.Cursor, data: tuple) -> None:
    matnr, bezeichnung, einheit, grenzwert, up_to, bestellt = data
    cursor.execute('''
        INSERT INTO Kleinstmaterial(MatNr,Bezeichnung,Einheit,Grenzwert,Auffüllen,bestellt)
        VALUES (?,?,?,?,?,?)''', (matnr, bezeichnung, einheit, grenzwert, up_to, bestellt))
    connection.commit()


def add_standardmaterial(connection: sqlite3.Connection,
                          cursor: sqlite3.Cursor, data: tuple) -> None:
    matnr, bezeichnung, einheit, grenzwert, up_to, bestellt = data
    cursor.execute('''
        INSERT INTO Standardmaterial(MatNr,Bezeichnung,Einheit,Grenzwert,Auffüllen,bestellt)
        VALUES (?,?,?,?,?,?)''', (matnr, bezeichnung, einheit, grenzwert, up_to, bestellt))
    connection.commit()


def add_bundle(connection: sqlite3.Connection, cursor: sqlite3.Cursor,
               matnr: int, bezeichnung: str, data: list) -> None:
    if not isinstance(data, list):
        print("Fehler: Datenformat muss Liste sein. z.B. [123,456]")
        return
    cursor.execute('''
        INSERT INTO Bundles(MatNr,Bezeichnung,Packungsinhalt) VALUES (?,?,?)''',
        (matnr, bezeichnung, str(data)))
    connection.commit()


def read_adresses_from_workorder_list(connection: sqlite3.Connection,
                                      cursor: sqlite3.Cursor) -> None:
    paths, _ = QFileDialog.getOpenFileNames(
        None, 'Excel-Datei auswählen', 'ExcelPYTHON.xlsm',
        'Excel (*.xlsx);;Alle Dateien (*.*)')
    if not paths:
        return
    for file in paths:
        excel_file = pd.read_excel(file)
        start_row  = 0
        all_entries_in_db  = cursor.execute(
            'SELECT SM_Nummer FROM Adresszuordnung').fetchall()
        all_known_sm_numbers = [r['SM_Nummer'] for r in all_entries_in_db]
        new_addresses = []
        seen          = set()
        for idx, line in excel_file.iterrows():
            if line['ID'] == 'Start':
                start_row = idx
        for _, line in excel_file.iloc[start_row:].iterrows():
            if pd.isnull(line['Was']):
                continue
            if 'OLT' in line['Was'] and 'Auskundung' not in line['Was']:
                if str(line['SM']) in all_known_sm_numbers or line['SM'] in seen:
                    continue
                new_addresses.append((str(line['SM']), line['VPSZ'], line['ORT']))
                seen.add(line['SM'])
    for smnr, vpsz, address in new_addresses:
        cursor.execute('''
            INSERT INTO Adresszuordnung (SM_Nummer, VPSZ, Adresse) VALUES (?,?,?)''',
            (smnr, vpsz, address))
        print(smnr, vpsz, address)
    connection.commit()
    print(f"Done. Added {len(new_addresses)} new addresses.")


def write_correction_data_from_yearly_inspection(
        connection: sqlite3.Connection, cursor: sqlite3.Cursor) -> None:
    paths, _ = QFileDialog.getOpenFileNames(
        None, 'Inventurdaten auswählen', 'Inventurdaten.xlsx',
        'Excel (*.xlsx);;Alle Dateien (*.*)')
    if not paths:
        return
    for file in paths:
        excel_file = pd.read_excel(file)
        for _, line in excel_file.iterrows():
            cursor.execute('''
                INSERT INTO Jahresinventur_Korrekturdaten (MatNr, Bezeichnung, Menge)
                VALUES (?,?,?)''',
                (line['MatNr'], line['Bezeichnung'], line['Differenz']))
    connection.commit()
